"""Read WC2026_SquadLists_WithPositions_AI.xlsx → regenerate
  data/players.json (with granular GK/CB/FB/MID/WIN/ST roles)
  data/teams.json (with the 6 categories)

Excel is the new source of truth, replacing the Wikipedia scrape.
"""
from __future__ import annotations
import json
import re
import unicodedata
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "WC2026_SquadLists_WithPositions_AI.xlsx"
OUT_PLAYERS = ROOT / "data" / "players.json"
OUT_TEAMS = ROOT / "data" / "teams.json"

# Excel name → canonical (Wikipedia/FIFA) name
ALIAS_TO_CANON = {
    "Bosnia":          "Bosnia and Herzegovina",
    "Cabo Verde":      "Cape Verde",
    "Congo DR":        "DR Congo",
    "Curacao":         "Curaçao",
    "Czechia":         "Czech Republic",
    "Côte D'Ivoire":   "Ivory Coast",
    "Korea":           "South Korea",
    "S Africa":        "South Africa",
    "USA":             "United States",
    "Uzbakistan":      "Uzbekistan",
}

# 3-letter FIFA code + flag emoji per canonical nation
META = {
    "Algeria":               ("ALG", "🇩🇿"),
    "Argentina":             ("ARG", "🇦🇷"),
    "Australia":             ("AUS", "🇦🇺"),
    "Austria":               ("AUT", "🇦🇹"),
    "Belgium":               ("BEL", "🇧🇪"),
    "Bosnia and Herzegovina":("BIH", "🇧🇦"),
    "Brazil":                ("BRA", "🇧🇷"),
    "Canada":                ("CAN", "🇨🇦"),
    "Cape Verde":            ("CPV", "🇨🇻"),
    "Colombia":              ("COL", "🇨🇴"),
    "Croatia":               ("CRO", "🇭🇷"),
    "Curaçao":               ("CUW", "🇨🇼"),
    "Czech Republic":        ("CZE", "🇨🇿"),
    "DR Congo":              ("COD", "🇨🇩"),
    "Ecuador":               ("ECU", "🇪🇨"),
    "Egypt":                 ("EGY", "🇪🇬"),
    "England":               ("ENG", "🏴\U000e0067\U000e0062\U000e0065\U000e006e\U000e0067\U000e007f"),
    "France":                ("FRA", "🇫🇷"),
    "Germany":               ("GER", "🇩🇪"),
    "Ghana":                 ("GHA", "🇬🇭"),
    "Haiti":                 ("HAI", "🇭🇹"),
    "Iran":                  ("IRN", "🇮🇷"),
    "Iraq":                  ("IRQ", "🇮🇶"),
    "Ivory Coast":           ("CIV", "🇨🇮"),
    "Japan":                 ("JPN", "🇯🇵"),
    "Jordan":                ("JOR", "🇯🇴"),
    "Mexico":                ("MEX", "🇲🇽"),
    "Morocco":               ("MAR", "🇲🇦"),
    "Netherlands":           ("NED", "🇳🇱"),
    "New Zealand":           ("NZL", "🇳🇿"),
    "Norway":                ("NOR", "🇳🇴"),
    "Panama":                ("PAN", "🇵🇦"),
    "Paraguay":              ("PAR", "🇵🇾"),
    "Portugal":              ("POR", "🇵🇹"),
    "Qatar":                 ("QAT", "🇶🇦"),
    "Saudi Arabia":          ("KSA", "🇸🇦"),
    "Scotland":              ("SCO", "🏴\U000e0067\U000e0062\U000e0073\U000e0063\U000e0074\U000e007f"),
    "Senegal":               ("SEN", "🇸🇳"),
    "South Africa":          ("RSA", "🇿🇦"),
    "South Korea":           ("KOR", "🇰🇷"),
    "Spain":                 ("ESP", "🇪🇸"),
    "Sweden":                ("SWE", "🇸🇪"),
    "Switzerland":           ("SUI", "🇨🇭"),
    "Tunisia":               ("TUN", "🇹🇳"),
    "Turkey":                ("TUR", "🇹🇷"),
    "United States":         ("USA", "🇺🇸"),
    "Uruguay":               ("URU", "🇺🇾"),
    "Uzbekistan":            ("UZB", "🇺🇿"),
}

# Arab League nations among qualifiers (for the ≥1 Arab constraint)
ARAB = {"Algeria","Egypt","Iraq","Jordan","Morocco","Qatar","Saudi Arabia","Tunisia"}


def canon(name: str) -> str:
    name = unicodedata.normalize("NFC", name.strip())
    return ALIAS_TO_CANON.get(name, name)


def parse_categories(wb) -> dict[str, int]:
    """Team Categories sheet → {canonical nation name: category 1..6}"""
    ws = wb["Team Categories"]
    out: dict[str, int] = {}
    for col_idx in range(1, 7):
        # First row is header "Category N"
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
            cell = row[0]
            if not cell:
                continue
            # "France - 1" → just want "France"
            name = re.split(r"\s*-\s*", cell, 1)[0].strip()
            out[canon(name)] = col_idx
    return out


def parse_players(wb) -> dict[str, list[dict]]:
    ws = wb["Players DB"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    by_nation: dict[str, list[dict]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[idx["Team"]]:
            continue
        nation = canon(row[idx["Team"]])
        roles: list[str] = []
        for col, label in [
            ("GK Yes (AI)", "GK"),
            ("CB Yes (AI)", "CB"),
            ("FB Yes (AI)", "FB"),
            ("Mid Yes (AI)", "CM"),
            ("Winger Yes (AI)", "WIN"),
            ("Striker Yes (AI)", "ST"),
        ]:
            if row[idx[col]] == "Yes":
                roles.append(label)
        if not roles:
            # fallback: derive from raw POS
            pos = row[idx["POS"]]
            roles = {"GK": ["GK"], "DF": ["CB"], "MF": ["CM"], "FW": ["ST"]}.get(pos, [])
        dob = row[idx["DOB"]]
        player = {
            "no": row[idx["Shirt #"]],
            "name": row[idx["PLAYER NAME"]],
            "shirt_name": row[idx["NAME ON SHIRT"]],
            "first": row[idx["FIRST NAME(S)"]],
            "last": row[idx["LAST NAME(S)"]],
            "club": row[idx["CLUB"]],
            "height": row[idx["HEIGHT (CM)"]],
            "dob": dob.strftime("%Y-%m-%d") if dob else None,
            "pos_raw": row[idx["POS"]],
            "roles": roles,
        }
        by_nation.setdefault(nation, []).append(player)
    # Sort within nation by shirt number
    for plist in by_nation.values():
        plist.sort(key=lambda p: (p["no"] is None, p["no"] or 0))
    return by_nation


def main() -> None:
    wb = openpyxl.load_workbook(SRC, data_only=True)
    categories = parse_categories(wb)
    players_by_nation = parse_players(wb)

    # Build teams.json
    teams = []
    for name in sorted(players_by_nation):
        if name not in META:
            print(f"!! Missing META for {name!r}")
            continue
        code, flag = META[name]
        teams.append({
            "name": name,
            "code": code,
            "flag": flag,
            "category": categories.get(name),
            "arab": name in ARAB,
        })
    missing_cat = [t["name"] for t in teams if t["category"] is None]
    if missing_cat:
        print(f"!! Nations without category: {missing_cat}")
    teams.sort(key=lambda t: (t["category"] or 99, t["name"]))
    OUT_TEAMS.write_text(json.dumps({"teams": teams}, ensure_ascii=False, indent=2))
    print(f"Wrote {len(teams)} teams → {OUT_TEAMS}")

    # Build players.json
    nations_out = []
    for nation in sorted(players_by_nation):
        nations_out.append({
            "name": nation,
            "category": categories.get(nation),
            "arab": nation in ARAB,
            "players": players_by_nation[nation],
        })
    OUT_PLAYERS.write_text(json.dumps({"nations": nations_out}, ensure_ascii=False, indent=2))
    total = sum(len(n["players"]) for n in nations_out)
    print(f"Wrote {len(nations_out)} nations, {total} players → {OUT_PLAYERS}")

    # Distribution check
    from collections import Counter
    role_count = Counter()
    multi = 0
    for n in nations_out:
        for p in n["players"]:
            if len(p["roles"]) > 1:
                multi += 1
            for r in p["roles"]:
                role_count[r] += 1
    print(f"Roles (with overlaps): {dict(role_count)} · multi-role: {multi}")
    by_cat = Counter(t["category"] for t in teams)
    print(f"Per category: {dict(by_cat)}")


if __name__ == "__main__":
    main()
